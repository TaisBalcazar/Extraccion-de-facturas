"""
Extractor para Categoría 4: Agua (Municipio de Loja) y Papel Bond.

Extrae datos de facturas de agua potable Y de reportes de entrega de papel bond.

──────────────────────────────────────────────────────────────────────────────
AGUA (Municipio de Loja)
  Campos: Año | Mes Número | Mes | Medidor | Categoría | Ubicación
          Lectura Anterior | Lectura actual | Consumo m3
          Agua Potable | Recolección de Basura | Costo Básico de Facturación
          Protección de microcuencas | Seguridad Ciudadana | Aportes Planes Maestros
          Alcantarillado | Interés - Recargo | Total facturado

──────────────────────────────────────────────────────────────────────────────
PAPEL BOND (Reporte de entrega Mega Santiago / Luxal Corp)
  Detecta tablas mensuales de resmas A4 y A3.
  Por cada mes calcula:
    total_resmas       = resmas_A4 + resmas_A3  (todas las marcas/columnas)
    peso_papel_kg      = total_resmas × 2.33
    peso_papel_toneladas = peso_papel_kg / 1000
  Guarda los resultados en el campo `periodos` (uno por mes) para que el
  front-end pueda expandirlos como filas individuales en la tabla de detalle.
"""

import re
import io
from datetime import date
from typing import Dict, Any, Optional, List

import pdfplumber

from app.services.ocr_extractor import OCRNotAvailableError

from app.services.schemas import Category4Schema
from app.services.extractors.base import BaseCategoryExtractor

# ── Helpers de hospedaje ──────────────────────────────────────────────────────

_MESES_ESP: Dict[str, int] = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "SETIEMBRE": 9, "SEPTIMBRE": 9,
    "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4,
    "MAY": 5, "JUN": 6, "JUL": 7, "AGO": 8,
    "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}

_MESES_NOMBRES: Dict[int, str] = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}


def _nombre_mes_hosp(num: int) -> str:
    return _MESES_NOMBRES.get(num, str(num))


def _calcular_noches(fecha_str: Any) -> int:
    """
    Calcula noches de hospedaje a partir del texto de una celda de fechas.

    Formatos soportados:
      "11-12 ENERO"          → 1 noche   (mismo mes: 12 - 11)
      "01-03 ABRIL"          → 2 noches  (mismo mes: 3 - 1)
      "30 ABRIL -01 MAYO"    → 1 noche   (cruce de mes)
      "18 enero"             → 1 noche   (fecha única)
      ""  / NaN / vacío      → 1 noche
    """
    if fecha_str is None:
        return 1
    s = str(fecha_str).strip()
    if s.lower() in ("", "nan", "none"):
        return 1

    s = s.upper()
    s = re.sub(r"/.*$", "", s).strip()        # quitar "/ TRANSFER" etc.
    s = re.sub(r"\b20\d{2}\b", "", s).strip() # quitar años "2025"

    # Rango cruzando mes: "DD MES -DD MES"
    cross = re.search(
        r"(\d{1,2})\s+([A-ZÁÉÍÓÚÑ]+)\s*[-–]\s*(\d{1,2})\s+([A-ZÁÉÍÓÚÑ]+)", s
    )
    if cross:
        d1, m1 = int(cross.group(1)), _MESES_ESP.get(cross.group(2), 0)
        d2, m2 = int(cross.group(3)), _MESES_ESP.get(cross.group(4), 0)
        if m1 and m2:
            try:
                year = 2025
                date1 = date(year, m1, d1)
                date2 = date(year, m2, d2)
                if date2 <= date1:
                    date2 = date(year + 1, m2, d2)
                return max(1, (date2 - date1).days)
            except (ValueError, OverflowError):
                pass
        return 1

    # Rango mismo mes: "DD-DD"
    same = re.search(r"(\d{1,2})\s*[-–]\s*(\d{1,2})", s)
    if same:
        return max(1, int(same.group(2)) - int(same.group(1)))

    return 1  # fecha única o texto no reconocido


def _detectar_columnas_excel(header_row: tuple) -> Dict[str, int]:
    cols: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        val = str(cell).upper().strip() if cell is not None else ""
        if val.startswith("MES") and "FECHA" not in val:
            cols.setdefault("mes", idx)
        elif "FECHA" in val:
            cols.setdefault("fechas", idx)
        elif "PROVEEDOR" in val or ("RAZ" in val and "SOCIAL" in val):
            cols.setdefault("proveedor", idx)
        elif "NOCHE" in val:
            cols.setdefault("noches", idx)
    return cols


class Category4Extractor(BaseCategoryExtractor):
    """Extractor para Categoría 4: detecta automáticamente agua o papel bond."""

    # ──────────────────────────────────────────────────────────
    # Constantes de meses (compartidas por ambos modos)
    # ──────────────────────────────────────────────────────────

    MESES = {
        "enero":      ("01_ENE", 1),  "febrero":   ("02_FEB", 2),
        "marzo":      ("03_MAR", 3),  "abril":     ("04_ABR", 4),
        "mayo":       ("05_MAY", 5),  "junio":     ("06_JUN", 6),
        "julio":      ("07_JUL", 7),  "agosto":    ("08_AGO", 8),
        "septiembre": ("09_SEP", 9),  "octubre":   ("10_OCT", 10),
        "noviembre":  ("11_NOV", 11), "diciembre": ("12_DIC", 12),
    }

    # Meses en mayúsculas → número (para parseo de tabla bond)
    MESES_BOND: Dict[str, int] = {
        "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
        "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
        "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
    }

    # Número → nombre capitalizado
    MESES_NOMBRE: Dict[int, str] = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }

    # Nombre del mes (es) → número de dos dígitos (para fecha)
    MESES_TEXTO_NUM: Dict[str, str] = {
        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
        "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
        "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
    }

    def __init__(self):
        super().__init__(Category4Schema())

    # ──────────────────────────────────────────────────────────
    # Utilidades numéricas (mismas que Category2)
    # ──────────────────────────────────────────────────────────

    @staticmethod
    def _to_float(raw_value: str) -> Optional[float]:
        normalized = (raw_value or "").strip().replace(" ", "")
        if not normalized:
            return None
        normalized = re.sub(r"(?<=\d)[oO](?=\d)", "0", normalized)
        normalized = re.sub(r"(?<=\d)[iIlL](?=\d)", "1", normalized)
        if "," in normalized and "." in normalized:
            if normalized.rfind(",") > normalized.rfind("."):
                normalized = normalized.replace(".", "").replace(",", ".")
            else:
                normalized = normalized.replace(",", "")
        elif "," in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None

    def _ef(self, patterns: list, text: str) -> Optional[float]:
        """extract_float: primer match válido entre los patrones dados."""
        for pattern in patterns:
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                v = self._to_float(m.group(1))
                if v is not None:
                    return v
        return None

    # ──────────────────────────────────────────────────────────
    # Lectura del PDF
    # ──────────────────────────────────────────────────────────

    @staticmethod
    def _read_pdf_text(file_content: bytes) -> str:
        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)

    # ──────────────────────────────────────────────────────────
    # Detección de tipo de documento
    # ──────────────────────────────────────────────────────────

    @staticmethod
    def _is_papel_bond_report(texto: str) -> bool:
        """
        Detecta si el PDF es un reporte de entrega de papel bond
        (Mega Santiago / Luxal Corp.) y NO una factura de agua.

        Criterios:
        - Contiene al menos 2 de las palabras clave de papel bond.
        - Contiene al menos 3 nombres de meses (tabla mensual).
        - NO contiene marcadores típicos de factura de agua.
        """
        texto_lower = texto.lower()

        bond_keywords = ["resma", "bond", "papel bond", "resmas", "a4", "a3"]
        bond_hits = sum(1 for kw in bond_keywords if kw in texto_lower)

        meses_hits = sum(
            1 for m in ["enero", "febrero", "marzo", "abril", "mayo"]
            if m in texto_lower
        )

        agua_markers = ["municipio de loja", "servicio de agua potable", "alcantarillado"]
        es_agua = any(m in texto_lower for m in agua_markers)

        return bond_hits >= 2 and meses_hits >= 3 and not es_agua

    # ──────────────────────────────────────────────────────────
    # Extracción de línea de item de la factura (modo agua)
    # ──────────────────────────────────────────────────────────

    def _extract_item_value(self, keyword_pattern: str, text: str) -> Optional[float]:
        """
        Extrae el valor total de una línea de item con formato:
          '1.00  DESCRIPCION  V.unit  V.total'
        """
        pattern = r"1\.00\s+" + keyword_pattern + r"[^\n]*?([0-9]+[.,][0-9]+)\s*$"
        m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if m:
            return self._to_float(m.group(1))

        pattern_loose = keyword_pattern + r"[^\n]*?([0-9]+[.,][0-9]+)"
        m = re.search(pattern_loose, text, re.IGNORECASE)
        if m:
            return self._to_float(m.group(1))

        return None

    # ──────────────────────────────────────────────────────────
    # Método principal (dispatcher)
    # ──────────────────────────────────────────────────────────

    def extract(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Detecta automáticamente el tipo de documento y delega al
        extractor correspondiente (agua, papel bond o hospedaje Excel).
        """
        if filename.lower().endswith((".xlsx", ".xls")):
            print(f"[CAT4 EXTRACTOR] Excel detectado -> hospedaje: {filename}")
            return self._extract_hospedaje_excel(file_content, filename)

        try:
            print(f"[CAT4 EXTRACTOR] Leyendo PDF: {filename}")

            try:
                texto = self.extract_pdf_text(file_content)
            except OCRNotAvailableError:
                # No degradar a una lectura sin OCR: el texto plano ya fue
                # insuficiente (por eso se necesitaba OCR) — fallaría igual
                # pero de forma silenciosa. Mejor rechazar con mensaje claro.
                raise
            except Exception:
                texto = self._read_pdf_text(file_content)

            print(f"[CAT4 EXTRACTOR] Texto extraído: {len(texto)} caracteres")

            if self._is_papel_bond_report(texto):
                print("[CAT4 EXTRACTOR] Detectado: REPORTE DE PAPEL BOND")
                return self._extract_papel_bond(texto, filename)
            else:
                print("[CAT4 EXTRACTOR] Detectado: FACTURA DE AGUA")
                return self._extract_agua(texto, filename)

        except OCRNotAvailableError:
            raise
        except Exception as exc:
            print(f"[CAT4 EXTRACTOR] Error: {exc}")
            return self.create_error_response(str(exc))

    # ══════════════════════════════════════════════════════════
    # MODO HOSPEDAJE — lectura desde Excel
    # ══════════════════════════════════════════════════════════

    def _extract_hospedaje_excel(self, content: bytes, filename: str) -> Dict[str, Any]:
        """
        Extrae pernoctaciones en hoteles desde un Excel con columnas:
          MESES | FECHAS DE HOSPEDAJE | Razón social del proveedor | Número de noches

        Calcula automáticamente las noches a partir del rango de fechas y
        las acumula por mes (columna MESES).
        """
        try:
            import openpyxl
        except ImportError:
            return self.create_error_response(
                "openpyxl no disponible. Instala: pip install openpyxl"
            )

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
        except Exception as exc:
            return self.create_error_response(f"Error al abrir Excel: {exc}")

        # Buscar fila de encabezados (debe contener "MES" y "FECHA")
        header_row_idx: Optional[int] = None
        cols: Dict[str, int] = {}

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [str(c).upper().strip() if c is not None else "" for c in row]
            if any("MES" in v for v in vals) and any("FECHA" in v for v in vals):
                header_row_idx = row_idx
                cols = _detectar_columnas_excel(tuple(vals))
                break

        if header_row_idx is None or "mes" not in cols or "fechas" not in cols:
            return self.create_error_response(
                "No se encontraron encabezados MESES y FECHAS DE HOSPEDAJE. "
                "Verifica que el Excel tenga esas columnas."
            )

        print(f"[CAT4 HOSPEDAJE] Encabezados en fila {header_row_idx} | cols={cols}")

        noches_por_mes: Dict[int, int] = {}
        detalle: list = []
        proveedor: Optional[str] = None
        mes_actual: Optional[int] = None
        año: Optional[int] = None

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            val_mes = row[cols["mes"]] if cols["mes"] < len(row) else None
            val_fechas = row[cols["fechas"]] if cols["fechas"] < len(row) else None
            val_prov = (
                row[cols["proveedor"]]
                if "proveedor" in cols and cols["proveedor"] < len(row)
                else None
            )

            mes_str = str(val_mes).strip() if val_mes is not None else ""
            fechas_str = str(val_fechas).strip() if val_fechas is not None else ""
            if not mes_str and not fechas_str:
                continue

            if mes_str:
                m = _MESES_ESP.get(mes_str.upper())
                if m:
                    mes_actual = m

            if mes_actual is None:
                continue

            if val_prov and not proveedor:
                proveedor = str(val_prov).strip()

            if val_fechas and año is None:
                m_year = re.search(r"\b(20\d{2})\b", str(val_fechas))
                if m_year:
                    año = int(m_year.group(1))

            noches = _calcular_noches(val_fechas)
            noches_por_mes[mes_actual] = noches_por_mes.get(mes_actual, 0) + noches
            detalle.append({
                "mes_numero": mes_actual,
                "mes_nombre": _nombre_mes_hosp(mes_actual),
                "fechas": fechas_str,
                "noches_calculadas": noches,
            })

        total_noches = sum(noches_por_mes.values())
        noches_por_mes_nombres = {
            _nombre_mes_hosp(m): n for m, n in sorted(noches_por_mes.items())
        }

        print(
            f"✅ [CAT4 HOSPEDAJE] {len(detalle)} registros | "
            f"{total_noches} noches | meses={list(noches_por_mes_nombres.keys())}"
        )

        datos = self.initialize_data_dict()
        datos.update({
            "filename": filename,
            "tipo_documento": "hospedaje",
            "tipo_servicio": "Hospedaje",
            "proveedor": proveedor,
            "year": año,
            "noches_por_mes": noches_por_mes_nombres,
            "noches_por_mes_numero": {str(m): n for m, n in sorted(noches_por_mes.items())},
            "total_noches": total_noches,
            "detalle_registros": detalle,
            "total_registros": len(detalle),
            "extraction_success": True,
        })
        return datos

    # ══════════════════════════════════════════════════════════
    # MODO AGUA — lógica original sin modificaciones
    # ══════════════════════════════════════════════════════════

    def _extract_agua(self, texto: str, filename: str) -> Dict[str, Any]:
        """Extrae datos de factura de agua (Municipio de Loja)."""
        datos = self.initialize_data_dict()
        datos["filename"] = filename
        datos["tipo_documento"] = "agua"

        # ── Número de factura ──────────────────────────────
        m = re.search(r"nro\.?\s+([0-9]{3}-[0-9]{3}-[0-9]+)", texto, re.IGNORECASE)
        if m:
            datos["factura_numero"] = m.group(1).strip()

        # ── Fecha de emisión ───────────────────────────────
        m = re.search(r"fecha\s+de\s+emisi[oó?]n[:\s]+([0-9]{2}-[0-9]{2}-[0-9]{4})", texto, re.IGNORECASE)
        if m:
            datos["fecha_emision"] = m.group(1).strip()

        # ── Período / Año / Mes ────────────────────────────
        m = re.search(r"correspondiente\s+a[:\s]+(\d{4})-([a-záéíóú]+)", texto, re.IGNORECASE)
        if m:
            datos["year"] = int(m.group(1))
            mes_lower = m.group(2).strip().lower()
            datos["mes_nombre"] = mes_lower.capitalize()
            if mes_lower in self.MESES:
                datos["mes_codigo"], datos["mes_numero"] = self.MESES[mes_lower]

        # ── Número de medidor ──────────────────────────────
        m = re.search(r"n[?ú]mero\s+de\s+medidor\s+([0-9]+)", texto, re.IGNORECASE)
        if m:
            datos["medidor"] = m.group(1).strip()

        # ── Estado del medidor ─────────────────────────────
        m = re.search(r"estado\s+medidor\s+([A-Za-záéíóúÁÉÍÓÚñÑ]+)", texto, re.IGNORECASE)
        if m:
            datos["estado_medidor"] = m.group(1).strip().capitalize()

        # ── Categoría del medidor ──────────────────────────
        # Acotado al bloque INFORMACION ADICIONAL para evitar capturar frases como
        # "Categoría de facturación" que podrían aparecer antes en el texto.
        _info_m = re.search(
            r"INFORMACI[O\?Ó]N\s+ADICIONAL(.+?)(?=\n\s*\n|\Z)", texto, re.IGNORECASE | re.DOTALL
        )
        _scope = _info_m.group(1) if _info_m else texto
        m = re.search(r"categor[?í]a\s+([A-Za-záéíóúÁÉÍÓÚ]+)", _scope, re.IGNORECASE)
        if m:
            datos["categoria"] = m.group(1).strip().capitalize()

        # ── Ubicación ──────────────────────────────────────
        m = re.search(r"ubicaci[oó?]n[:\s]+([^\n]+)", texto, re.IGNORECASE)
        if m:
            datos["ubicacion"] = m.group(1).strip()

        # ── Lecturas ───────────────────────────────────────
        datos["lectura_anterior"] = self._ef(
            [r"lectura\s+anterior\s+([0-9][0-9\.,]*)"],
            texto,
        )
        datos["lectura_actual"] = self._ef(
            [r"lectura\s+actual\s+([0-9][0-9\.,]*)"],
            texto,
        )

        # ── Consumo m³ ─────────────────────────────────────
        datos["consumo_m3"] = self._ef(
            [
                r"consumo\s*\(m3\)\s+([0-9][0-9\.,]*)",
                r"consumo\s*\(?m[³3]\)?\s+([0-9][0-9\.,]*)",
            ],
            texto,
        )
        if datos.get("consumo_m3") is None:
            la = datos.get("lectura_actual")
            lp = datos.get("lectura_anterior")
            if la is not None and lp is not None and la >= lp:
                datos["consumo_m3"] = round(la - lp, 3)
                print(f"ℹ️ [CAT4 AGUA] consumo_m3 calculado por diferencia: {datos['consumo_m3']}")

        # ── Ítems monetarios ───────────────────────────────
        datos["agua_potable"] = self._extract_item_value(
            r"SERVICIO\s+DE\s+AGUA\s+POTABLE", texto
        )
        datos["alcantarillado"] = self._extract_item_value(
            r"ALCANTARILLADO", texto
        )
        datos["aportes_planes_maestros"] = self._extract_item_value(
            r"APORTES?\s+PLANES?\s+MAESTROS?", texto
        )
        datos["seguridad_ciudadana"] = self._extract_item_value(
            r"SEGURIDAD\s+CIUDADANA", texto
        )
        datos["proteccion_microcuencas"] = self._extract_item_value(
            r"PROTECCION\s+MICROCUENCAS?", texto
        )
        datos["costo_basico_facturacion"] = self._extract_item_value(
            r"COSTO\s+BASICO\s+DE\s+FACTURACION", texto
        )
        datos["recoleccion_basura"] = self._extract_item_value(
            r"RECOLECCI[OÓ?N]+\s+DE\s+BASURA", texto
        )

        # ── Interés / Recargo ──────────────────────────────
        datos["interes_recargo"] = self._ef(
            [
                r"inter[eé]s\s*[/\-]?\s*recargo\s*[:\-]?\s*\$?\s*([0-9][0-9\.,]*)",
                r"recargo\s*[:\-]?\s*\$?\s*([0-9][0-9\.,]*)",
            ],
            texto,
        )
        if datos.get("interes_recargo") is None:
            datos["interes_recargo"] = 0.0

        # ── Total facturado ────────────────────────────────
        m_total = re.search(r"\bTotal\s+([0-9]+[.,][0-9]+)\s*$", texto, re.IGNORECASE | re.MULTILINE)
        if m_total:
            datos["total_facturado"] = self._to_float(m_total.group(1))

        if datos.get("total_facturado") is None:
            componentes = [
                datos.get("agua_potable"),
                datos.get("alcantarillado"),
                datos.get("aportes_planes_maestros"),
                datos.get("seguridad_ciudadana"),
                datos.get("proteccion_microcuencas"),
                datos.get("costo_basico_facturacion"),
                datos.get("recoleccion_basura"),
            ]
            if all(c is not None for c in componentes):
                datos["total_facturado"] = round(
                    sum(float(c) for c in componentes) + float(datos.get("interes_recargo") or 0),
                    2,
                )
                print(f"ℹ️ [CAT4 AGUA] total_facturado calculado por suma: {datos['total_facturado']}")

        # ── Diagnóstico ────────────────────────────────────
        campos_clave = ["consumo_m3", "agua_potable", "alcantarillado", "total_facturado"]
        faltantes = [c for c in campos_clave if datos.get(c) is None]
        if faltantes:
            lineas = [
                l.strip() for l in texto.splitlines()
                if l.strip() and any(k in l.lower() for k in ["agua", "consumo", "total", "m3", "basura"])
            ]
            print(f"[CAT4 AGUA] Campos no detectados: {faltantes}. Líneas candidatas: {lineas[:8]}")

        datos["extraction_success"] = True
        datos["texto_completo"] = texto[:500]

        print(f"[CAT4 AGUA] Extracción exitosa")
        return datos

    # ══════════════════════════════════════════════════════════
    # MODO PAPEL BOND — extracción de tabla mensual
    # ══════════════════════════════════════════════════════════

    def _extract_papel_bond(self, texto: str, filename: str) -> Dict[str, Any]:
        """
        Extrae el reporte mensual de resmas de papel bond (A4 + A3).

        Para cada mes con datos:
          total_resmas       = suma de todas las columnas numéricas de esa fila
          peso_papel_kg      = total_resmas × 2.3
          peso_papel_toneladas = peso_papel_kg / 1000

        Los datos mensuales se guardan en el campo `periodos` para que el
        front-end los expanda como filas individuales en la tabla de detalle.
        """
        datos = self.initialize_data_dict()
        datos["filename"] = filename
        datos["tipo_documento"] = "papel_bond"

        # ── Año ──────────────────────────────────────────────
        m_year = re.search(r'\b(20\d{2})\b', texto)
        if m_year:
            datos["year"] = int(m_year.group(1))

        # ── Fecha de emisión (ej: "14 de enero de 2025") ────
        m_fecha = re.search(
            r'(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|'
            r'septiembre|octubre|noviembre|diciembre)\s+de\s+(\d{4})',
            texto, re.IGNORECASE
        )
        if m_fecha:
            dia = m_fecha.group(1).zfill(2)
            mes_str = m_fecha.group(2).lower()
            anio_str = m_fecha.group(3)
            mes_num_str = self.MESES_TEXTO_NUM.get(mes_str, "01")
            # Formato ISO para que el campo `year` en Firestore se calcule bien
            datos["fecha_emision"] = f"{anio_str}-{mes_num_str}-{dia}"

        # ── Parsear tabla mensual ─────────────────────────────
        # Se busca cada línea que EMPIECE con el nombre de un mes.
        # Todos los enteros encontrados en esa línea (después del mes) se suman.
        # Esto funciona tanto para:
        #   "ENERO 40"        (una columna)
        #   "ENERO 40 0"      (dos columnas)
        #   "ABRIL 10 240"    (dos columnas con valores)
        periodos: List[Dict[str, Any]] = []
        seen_months: set = set()

        for linea in texto.splitlines():
            linea_strip = linea.strip()
            linea_upper = linea_strip.upper()

            for mes_nombre_upper, mes_num in self.MESES_BOND.items():
                if linea_upper.startswith(mes_nombre_upper) and mes_num not in seen_months:
                    seen_months.add(mes_num)

                    # Texto tras el nombre del mes
                    resto = linea_strip[len(mes_nombre_upper):].strip()

                    # Extraer enteros simples (los valores de resmas nunca son decimales)
                    # Usamos \b para no partir "1.582" en "1" y "582" — al ser un entero
                    # de mes nunca tendrá miles separados con punto, así que \d+ basta.
                    nums = [int(n) for n in re.findall(r'\d+', resto)]
                    total_resmas = sum(nums)

                    peso_kg = round(total_resmas * 2.33, 3)
                    peso_ton = round(peso_kg / 1000, 6)

                    periodos.append({
                        "mes_numero": mes_num,
                        "mes_nombre": self.MESES_NOMBRE[mes_num],
                        "total_resmas": total_resmas,
                        "peso_papel_kg": peso_kg,
                        "peso_papel_toneladas": peso_ton,
                        "year": datos.get("year"),
                    })
                    break

        # Ordenar cronológicamente
        periodos.sort(key=lambda p: p["mes_numero"])

        # ── Totales anuales ────────────────────────────────────
        total_resmas_anual = sum(p["total_resmas"] for p in periodos)
        peso_kg_anual = round(total_resmas_anual * 2.33, 3)
        peso_ton_anual = round(peso_kg_anual / 1000, 6)

        datos["periodos"] = periodos
        datos["total_resmas_anual"] = total_resmas_anual
        datos["peso_papel_kg_anual"] = peso_kg_anual
        datos["peso_papel_ton_anual"] = peso_ton_anual
        datos["extraction_success"] = True
        datos["texto_completo"] = texto[:500]

        print(
            f"✅ [CAT4 BOND] Extraídos {len(periodos)} meses. "
            f"Total anual: {total_resmas_anual} resmas | "
            f"{peso_kg_anual} kg | {peso_ton_anual} ton"
        )
        return datos
