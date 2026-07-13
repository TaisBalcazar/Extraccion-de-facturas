"""
Extractor para Categoría 3: Vuelos (nacionales e internacionales).

Lee un Excel con registros de tickets aéreos. Calcula la distancia (pas×km)
por ruta usando airportsdata + fórmula Haversine, y agrega por mes separando
vuelos nacionales e internacionales.

Formato de ruta esperado:
  - 'LOH-UIO-LOH'  → ida y vuelta  (Loja → Quito → Loja)
  - 'LOH-UIO'      → solo ida       (Loja → Quito)
El tipo se determina comprobando si todos los aeropuertos de la ruta están en Ecuador.
"""

import io
from math import asin, cos, radians, sin, sqrt
from typing import Any, Dict, Optional, Tuple

import openpyxl

from app.services.schemas import Category3Schema
from app.services.extractors.base import BaseCategoryExtractor

# ── Mapas de meses ───────────────────────────────────────────

MESES_MAP: Dict[str, int] = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
    'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
}

MESES_NOMBRE: Dict[int, str] = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}

# ── Base de datos de aeropuertos (carga única) ────────────────

_airports_db: Optional[Dict] = None


def _get_airports() -> Dict:
    global _airports_db
    if _airports_db is None:
        import airportsdata
        _airports_db = airportsdata.load('IATA')
    return _airports_db


# ── Cálculo de distancia ──────────────────────────────────────

def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> int:
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat, dlon = lat2 - lat1, lon2 - lon1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    return round(2 * 6371 * asin(sqrt(a)))


def _calcular_ruta(ruta: str) -> Tuple[int, str]:
    """
    Retorna (km_totales, tipo) para una cadena de ruta.

    Ejemplos:
      'LOH-UIO-LOH' → (884, 'nacional')
      'UIO-LOH'     → (442, 'nacional')
      'UIO-MIA'     → (3032, 'internacional')
    """
    airports = _get_airports()
    codes = [c.strip().upper() for c in ruta.split('-') if c.strip()]

    if len(codes) < 2:
        return 0, 'desconocido'

    total_km = 0
    paises: set = set()

    for i in range(len(codes) - 1):
        orig = airports.get(codes[i])
        dest = airports.get(codes[i + 1])
        if orig and dest:
            total_km += _haversine_km(orig['lat'], orig['lon'], dest['lat'], dest['lon'])
            paises.update([orig['country'], dest['country']])

    if total_km == 0:
        return 0, 'desconocido'

    tipo = 'nacional' if paises == {'EC'} else 'internacional'
    return total_km, tipo


# ── Detección de encabezados ──────────────────────────────────

def _find_header_row(ws) -> Tuple[Optional[int], Dict[str, int]]:
    """
    Busca la fila que contiene 'MESES' y 'RUTA' como encabezados.
    Retorna (número_de_fila, mapa_columna→índice).
    """
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [str(c).strip().upper() if c is not None else '' for c in row]
        if 'MESES' in normalized and 'RUTA' in normalized:
            col_map = {val: idx for idx, val in enumerate(normalized) if val}
            return row_idx, col_map
    return None, {}


# ── Extractor principal ───────────────────────────────────────

class Category3Extractor(BaseCategoryExtractor):
    """Extractor de vuelos nacionales e internacionales desde Excel."""

    def __init__(self):
        super().__init__(Category3Schema())

    def extract(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        print(f"[CAT3 EXTRACTOR] Procesando vuelos: {filename}")
        datos = self.initialize_data_dict()
        datos['filename'] = filename
        datos['tipo_servicio'] = 'Vuelos'

        try:
            datos.update(self._procesar_excel(file_content))
            datos['extraction_success'] = True
        except Exception as e:
            print(f"[CAT3 EXTRACTOR] Error: {e}")
            datos['extraction_success'] = False
            datos['error'] = str(e)

        return datos

    def _procesar_excel(self, file_content: bytes) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active

        header_row, col_map = _find_header_row(ws)
        if header_row is None:
            raise ValueError(
                "No se encontró la fila de encabezados. "
                "El Excel debe tener columnas 'MESES' y 'RUTA'."
            )

        col_mes = col_map.get('MESES')
        col_ruta = col_map.get('RUTA')
        if col_mes is None or col_ruta is None:
            raise ValueError(
                f"Columnas requeridas no encontradas. "
                f"Disponibles: {list(col_map.keys())}"
            )

        # N° DE TIKET puede tener variaciones en el nombre
        col_ticket = next(
            (col_map[k] for k in col_map if 'TIKET' in k or 'TICKET' in k),
            None
        )
        # Columna AÑO explícita en el Excel
        col_anio = next(
            (col_map[k] for k in col_map if k in ('AÑO', 'ANO', 'YEAR', 'AÑO DE VIAJE')),
            None
        )

        km_nacional: Dict[int, int] = {}
        km_internacional: Dict[int, int] = {}
        # (mes_num, ruta, tipo) → {cantidad, km_por_viaje}
        rutas_conteo: Dict[tuple, dict] = {}
        detalle = []
        anio: Optional[int] = None
        ultimo_mes_raw = None  # para manejar celdas combinadas en MESES

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(c is None for c in row):
                continue

            # Mes: reutiliza el último valor si la celda está vacía (celda combinada)
            mes_raw = row[col_mes] if col_mes < len(row) else None
            if mes_raw is not None:
                ultimo_mes_raw = mes_raw
            elif ultimo_mes_raw is not None:
                mes_raw = ultimo_mes_raw

            ruta_raw = row[col_ruta] if col_ruta < len(row) else None

            if not mes_raw or not ruta_raw:
                continue

            mes_str = str(mes_raw).strip().upper().split()[0]
            mes_num = MESES_MAP.get(mes_str)
            if not mes_num:
                continue

            ruta = str(ruta_raw).strip()
            km, tipo = _calcular_ruta(ruta)
            if km == 0:
                print(f"[CAT3] Ruta sin km calculable: '{ruta}'")
                continue

            # Año: columna AÑO explícita → objeto fecha → texto "DD M YYYY"
            if anio is None:
                if col_anio is not None and col_anio < len(row) and row[col_anio]:
                    try:
                        anio = int(row[col_anio])
                    except (TypeError, ValueError):
                        pass
                if anio is None:
                    for val in row:
                        if hasattr(val, 'year') and val.year and val.year > 2000:
                            anio = val.year
                            break
                        if isinstance(val, str):
                            parts = val.strip().split()
                            if len(parts) == 3:
                                try:
                                    candidate = int(parts[2])
                                    if candidate > 2000:
                                        anio = candidate
                                        break
                                except (ValueError, IndexError):
                                    pass

            n_ticket = str(row[col_ticket]).strip() if col_ticket is not None and row[col_ticket] else None

            if tipo == 'nacional':
                km_nacional[mes_num] = km_nacional.get(mes_num, 0) + km
            else:
                km_internacional[mes_num] = km_internacional.get(mes_num, 0) + km

            # Conteo por ruta dentro del mes
            ruta_key = (mes_num, ruta, tipo)
            if ruta_key not in rutas_conteo:
                rutas_conteo[ruta_key] = {'cantidad': 0, 'km_por_viaje': km}
            rutas_conteo[ruta_key]['cantidad'] += 1

            detalle.append({
                'mes': mes_num,
                'ruta': ruta,
                'n_ticket': n_ticket,
            })

        rutas_por_mes = [
            {
                'mes_num':      k[0],
                'mes_nombre':   MESES_NOMBRE[k[0]],
                'ruta':         k[1],
                'tipo':         k[2],
                'km_por_viaje': v['km_por_viaje'],
                'cantidad':     v['cantidad'],
                'km_total':     v['km_por_viaje'] * v['cantidad'],
            }
            for k, v in sorted(rutas_conteo.items())
        ]

        return {
            'anio': anio,
            'km_nacional_por_mes':      {str(k): v for k, v in sorted(km_nacional.items())},
            'km_internacional_por_mes': {str(k): v for k, v in sorted(km_internacional.items())},
            'rutas_por_mes':            rutas_por_mes,
            'detalle_registros':        detalle,
            'total_registros':          len(detalle),
        }
