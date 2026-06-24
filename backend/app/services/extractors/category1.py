"""
Extractor para Categoría 1: Reporte de Consumo de Combustible.

Procesa el Excel anual de combustible (REPORTE DE CONSUMO COMBUSTIBLE XXXX).
Detecta tipos de combustible, calcula totales y galones por vehículo y grupo.
"""

import re
import io
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

from app.services.schemas import Category1Schema
from app.services.extractors.base import BaseCategoryExtractor


# ─────────────────────────────────────────────────────────────────────────────
# Precios por defecto (USD/galón). Se sobreescriben desde Firestore en runtime.
# ─────────────────────────────────────────────────────────────────────────────

PRECIOS_COMBUSTIBLE_DEFAULT: Dict[str, float] = {
    "diesel": 2.10,
    "gasolina_ecopais": 2.70,
    "gasolina_movil": 3.60,
}

# ─────────────────────────────────────────────────────────────────────────────
# Catálogos de mapeo
# ─────────────────────────────────────────────────────────────────────────────

# Texto de columna "Tipo de combustible" (normalizado) → clave interna
_TIPO_COMBUSTIBLE_TEXTO: Dict[str, str] = {
    "DIESEL CONSUMIDO EN BUSES BUSETAS Y VEHICULOS": "diesel_buses",
    "DIESEL CONSUMIDO EN GENERADORES": "diesel_generadores",
    "GASOLINA ECOPAIS CONSUMIDA EN VEHICULOS": "gasolina_ecopais",
    "GASOLINA MOVIL CONSUMIDA EN VEHICULOS": "gasolina_movil",
}

# Clave interna → clave de precio en el dict de precios
_CLAVE_A_PRECIO: Dict[str, str] = {
    "diesel_buses": "diesel",
    "diesel_generadores": "diesel",
    "gasolina_ecopais": "gasolina_ecopais",
    "gasolina_movil": "gasolina_movil",
}

# Clave interna → etiqueta legible para mostrar en UI
_CLAVE_LABEL: Dict[str, str] = {
    "diesel_buses": "DIESEL CONSUMIDO EN BUSES BUSETAS Y VEHICULOS",
    "diesel_generadores": "DIESEL CONSUMIDO EN GENERADORES",
    "gasolina_ecopais": "GASOLINA ECOPAIS CONSUMIDA EN VEHÍCULOS",
    "gasolina_movil": "GASOLINA MÓVIL CONSUMIDA EN VEHÍCULOS",
}

# TIPO DE VEHICULO → clave de combustible (fallback cuando no hay columna explícita)
_TIPO_VEHICULO_MAPA: Dict[str, str] = {
    "BUS": "diesel_buses",
    "CAMION": "diesel_buses",
    "GENERADOR": "diesel_generadores",
    # JEEP, CAMIONETA, FURGONETA → gasolina_ecopais por defecto
}

# Vehículos con asignación especial (nombre normalizado → clave)
# Añadir aquí vehículos cuyo tipo de vehículo no alcanza para determinar el combustible.
_CATALOGO_ESPECIAL: Dict[str, str] = {
    "BLAZER CANCILLERIA": "gasolina_movil",
}

# Abreviaturas de meses en español → (número, nombre completo)
_MESES_ABBR: Dict[str, Tuple[int, str]] = {
    "ENE": (1, "enero"),  "FEB": (2, "febrero"),  "MAR": (3, "marzo"),
    "ABR": (4, "abril"),  "MAY": (5, "mayo"),      "JUN": (6, "junio"),
    "JUL": (7, "julio"),  "AGO": (8, "agosto"),    "SEP": (9, "septiembre"),
    "OCT": (10, "octubre"), "NOV": (11, "noviembre"), "DIC": (12, "diciembre"),
}

_MES_NUM_A_ABBR: Dict[int, str] = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DIC",
}


# ─────────────────────────────────────────────────────────────────────────────
# Utilidades internas
# ─────────────────────────────────────────────────────────────────────────────

def _norm(texto: Any) -> str:
    """Normaliza a mayúsculas sin acentos ni diacríticos."""
    t = unicodedata.normalize("NFD", str(texto or "").upper().strip())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def _to_float(val: Any) -> float:
    """Convierte celda de Excel a float; retorna 0.0 si no se puede."""
    if val is None:
        return 0.0
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# Detección de línea de combustible en facturas PDF
# ─────────────────────────────────────────────────────────────────────────────

def _detectar_linea_combustible(texto: str) -> Tuple[Optional[float], Optional[str], Optional[float], Optional[float]]:
    """
    Busca la fila de combustible en el detalle de la factura PDF.

    Cubre tres tipos de factura:
    - Facturas simples (DIESEL PREMIUM, Extra): un solo ítem de combustible.
    - Facturas mixtas (generador): múltiples ítems; solo se extrae la fila
      "GALON DE COMBUSTIBLE", ignorando mantenimiento/transporte.

    Returns:
        (cantidad, descripcion, precio_unitario, total_linea)
        Cualquier valor puede ser None si no se encuentra.
    """

    def _trio(texto_local: str) -> Tuple[Optional[float], Optional[float]]:
        """Busca (precio_unit, total) como el par [número] [descuento] [número]."""
        m = re.search(r"([\d.,]+)\s+\$?[\d.,]+\s+\$?([\d.,]+)", texto_local)
        if m:
            return _to_float(m.group(1)), _to_float(m.group(2))
        return None, None

    # 1. GALON [DE COMBUSTIBLE] — facturas de generadores u otros servicios mixtos
    #    Ej: "50.00 GALON DE COMBUSTIBLE 1.75 $0.00 $87.50"
    m = re.search(
        r"(\d+[.,]\d*)\s+GALON(?:ES)?(?:\s+DE\s+COMBUSTIBLE)?",
        texto, re.IGNORECASE
    )
    if m:
        cantidad = _to_float(m.group(1))
        precio_unit, total = _trio(texto[m.end(): m.end() + 200])
        return cantidad, "GALON DE COMBUSTIBLE", precio_unit, total

    # 2. DIESEL [PREMIUM / EXTRA / NORMAL / CORRIENTE]
    #    Ej: "17.446 DIESEL PREMIUM 1.562609 0.00 27.26"
    m = re.search(
        r"(\d+[.,]\d{2,})\s+(DIESEL(?:\s+(?:PREMIUM|EXTRA|NORMAL|CORRIENTE))?)\b",
        texto, re.IGNORECASE
    )
    if m:
        cantidad = _to_float(m.group(1))
        desc = m.group(2).strip().upper()
        precio_unit, total = _trio(texto[m.end(): m.end() + 300])
        return cantidad, desc, precio_unit, total

    # 3. GASOLINA [EXTRA / SUPER / REGULAR / CORRIENTE]
    m = re.search(
        r"(\d+[.,]\d{2,})\s+(GASOLINA(?:\s+(?:EXTRA|SUPER|REGULAR|CORRIENTE))?)\b",
        texto, re.IGNORECASE
    )
    if m:
        cantidad = _to_float(m.group(1))
        desc = m.group(2).strip().upper()
        precio_unit, total = _trio(texto[m.end(): m.end() + 300])
        return cantidad, desc, precio_unit, total

    # 4. EXTRA / SUPER / PREMIUM / GLP como nombre del producto
    #    Ej: "6.87 Extra Unidad Litros ... 2.53 0.00 17.39"
    m = re.search(r"(\d+[.,]\d{2,})\s+(EXTRA|SUPER|PREMIUM|GLP)\b", texto, re.IGNORECASE)
    if m:
        cantidad = _to_float(m.group(1))
        desc = m.group(2).strip().upper()
        precio_unit, total = _trio(texto[m.end(): m.end() + 300])
        return cantidad, desc, precio_unit, total

    return None, None, None, None


# ─────────────────────────────────────────────────────────────────────────────
# Extractor
# ─────────────────────────────────────────────────────────────────────────────

class Category1Extractor(BaseCategoryExtractor):
    """Extractor para Categoría 1: Reporte anual de consumo de combustible (Excel)."""

    def __init__(self, precios: Optional[Dict[str, float]] = None):
        super().__init__(Category1Schema())
        # Combinar defaults con overrides de Firestore
        self.precios: Dict[str, float] = {**PRECIOS_COMBUSTIBLE_DEFAULT, **(precios or {})}

    # ── Interfaz pública ──────────────────────────────────────────────────────

    def extract(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        if filename.lower().endswith((".xlsx", ".xls")):
            return self._extract_excel(file_content, filename)
        if filename.lower().endswith(".pdf"):
            return self._extract_pdf(file_content, filename)
        return self.create_error_response(
            "Categoría 1 acepta Excel (.xlsx) para reportes anuales (Zona Loja) "
            "o PDF para facturas individuales de combustible."
        )

    # ── Detección de tipo de combustible (cascada) ────────────────────────────

    def _detectar_tipo(
        self,
        col_combustible: Optional[str],
        tipo_vehiculo: Optional[str],
        nombre_vehiculo: Optional[str],
    ) -> str:
        # 1) Columna "Tipo de combustible" explícita
        if col_combustible:
            clave = _TIPO_COMBUSTIBLE_TEXTO.get(_norm(col_combustible))
            if clave:
                return clave

        # 2) Catálogo de vehículos especiales (nombre normalizado, solo alfanumérico)
        if nombre_vehiculo:
            nombre_clean = re.sub(r"[^A-Z0-9 ]", "", _norm(nombre_vehiculo)).strip()
            if nombre_clean in _CATALOGO_ESPECIAL:
                return _CATALOGO_ESPECIAL[nombre_clean]

        # 3) Columna TIPO DE VEHICULO
        if tipo_vehiculo:
            return _TIPO_VEHICULO_MAPA.get(_norm(tipo_vehiculo), "gasolina_ecopais")

        return "gasolina_ecopais"

    # ── Detección de columnas de meses en el encabezado ──────────────────────

    @staticmethod
    def _detectar_cols_mes(header_row: tuple) -> Dict[str, int]:
        """Retorna {nombre_mes: índice_columna} para las columnas ene-dic."""
        cols: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            # Celda con fecha real (datetime/date)
            if hasattr(cell, "month"):
                abbr = _MES_NUM_A_ABBR.get(cell.month)
                if abbr and abbr in _MESES_ABBR:
                    _, nombre = _MESES_ABBR[abbr]
                    cols.setdefault(nombre, idx)
                continue
            # Celda con texto: "ene.-25", "ENE-25", "ENE 25", "ENE", etc.
            val = _norm(cell)
            m = re.match(r"([A-Z]{3})[\.\-\s]*\d{0,2}$", val.replace(" ", "")) or re.match(r"([A-Z]{3})", val)
            if m:
                abbr = m.group(1)
                if abbr in _MESES_ABBR:
                    _, nombre = _MESES_ABBR[abbr]
                    cols.setdefault(nombre, idx)
        return cols

    @staticmethod
    def _detectar_anio(header_row: tuple) -> Optional[int]:
        for cell in header_row:
            if cell is None:
                continue
            if hasattr(cell, "year"):
                return cell.year
            m = re.search(r"[\.\-\s](\d{2})$", str(cell).strip())
            if m:
                return 2000 + int(m.group(1))
        return None

    # ── Detección de columnas de control ─────────────────────────────────────

    @staticmethod
    def _detectar_cols_control(header_row: tuple) -> Dict[str, Optional[int]]:
        """Mapea columnas ITM, VEHICULO, TIPO_VEHICULO, TIPO_COMBUSTIBLE."""
        cols: Dict[str, Optional[int]] = {
            "itm": None, "vehiculo": None,
            "tipo_vehiculo": None, "tipo_combustible": None,
        }
        for idx, cell in enumerate(header_row):
            v = _norm(cell)
            if cols["itm"] is None and ("ITM" in v or v == "ITEM"):
                cols["itm"] = idx
            elif cols["vehiculo"] is None and "VEHICULO" in v and "TIPO" not in v:
                cols["vehiculo"] = idx
            elif cols["tipo_vehiculo"] is None and "TIPO" in v and "VEHICULO" in v:
                cols["tipo_vehiculo"] = idx
            elif cols["tipo_combustible"] is None and "COMBUSTIBLE" in v and "TIPO" in v:
                cols["tipo_combustible"] = idx
        return cols

    # ── Extracción principal ──────────────────────────────────────────────────

    def _extract_excel(self, content: bytes, filename: str) -> Dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            return self.create_error_response(
                "openpyxl no disponible. Instala: pip install openpyxl"
            )

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
        except Exception as exc:
            return self.create_error_response(f"Error al abrir Excel: {exc}")

        # ── Buscar fila de encabezados (contiene ITM y VEHICULO) ─────────────
        header_row_idx: Optional[int] = None
        header_row: tuple = ()

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [_norm(c) for c in row]
            if any("ITM" in v or v == "ITEM" for v in vals) and any("VEHICULO" in v for v in vals):
                header_row_idx = row_idx
                header_row = tuple(row)
                break

        if not header_row_idx:
            return self.create_error_response(
                "No se encontró la fila de encabezados (ITM + VEHICULO). "
                "Verifica que el archivo sea el reporte de combustible."
            )

        cols_mes = self._detectar_cols_mes(header_row)
        if not cols_mes:
            return self.create_error_response(
                "No se detectaron columnas de meses (ENE-DIC) en el encabezado."
            )

        anio = self._detectar_anio(header_row)
        ctrl = self._detectar_cols_control(header_row)

        print(
            f"📊 [CAT1] Encabezados fila {header_row_idx} | año={anio} | "
            f"meses={list(cols_mes.keys())} | "
            f"col_tipo_comb={'sí' if ctrl['tipo_combustible'] is not None else 'no (auto-detección)'}"
        )

        # ── Procesar filas de datos ───────────────────────────────────────────
        def _get(row: tuple, key: str) -> Any:
            idx = ctrl.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        vehiculos: List[Dict] = []

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(c is None for c in row):
                continue

            val_veh = _get(row, "vehiculo")
            if not val_veh or not str(val_veh).strip():
                continue

            # Saltar fila de totales generales
            if any(kw in _norm(val_veh) for kw in ("TOTAL", "VALORES")):
                continue

            # ITM debe ser numérico → filtra filas de encabezado repetido
            val_itm = _get(row, "itm")
            try:
                itm_num = int(float(str(val_itm))) if val_itm is not None else None
            except (ValueError, TypeError):
                continue

            val_tipo_veh = _get(row, "tipo_vehiculo")
            val_tipo_comb = _get(row, "tipo_combustible")

            # Valores mensuales y total calculado
            meses: Dict[str, float] = {}
            for nombre_mes, idx_mes in cols_mes.items():
                meses[nombre_mes] = _to_float(row[idx_mes] if idx_mes < len(row) else None)

            total_usd = round(sum(meses.values()), 2)

            # Tipo de combustible y cálculo de galones
            tipo_clave = self._detectar_tipo(
                str(val_tipo_comb) if val_tipo_comb else None,
                str(val_tipo_veh) if val_tipo_veh else None,
                str(val_veh),
            )
            precio_key = _CLAVE_A_PRECIO.get(tipo_clave, "diesel")
            precio_galon = self.precios.get(precio_key) or PRECIOS_COMBUSTIBLE_DEFAULT["diesel"]
            galones = round(total_usd / precio_galon, 2) if precio_galon > 0 else 0.0

            vehiculos.append({
                "itm": itm_num,
                "vehiculo": str(val_veh).strip(),
                "tipo_vehiculo": str(val_tipo_veh).strip() if val_tipo_veh else None,
                "tipo_combustible_label": _CLAVE_LABEL.get(tipo_clave, tipo_clave),
                "tipo_combustible_clave": tipo_clave,
                "meses": meses,
                "total_usd": total_usd,
                "galones": galones,
            })

        if not vehiculos:
            return self.create_error_response(
                "No se encontraron filas de vehículos. "
                "Verifica que el Excel tenga datos debajo del encabezado ITM/VEHICULO."
            )

        # ── Mapeo nombre de mes → número para ordenar ─────────────────────────
        _mes_a_num: Dict[str, int] = {nombre: num for _, (num, nombre) in _MESES_ABBR.items()}

        # ── Agregar totales mensuales por tipo de combustible ─────────────────
        # Para cada grupo: acumular el total USD de todos los vehículos, mes a mes.
        # Los galones se calculan por mes: total_mes_usd / precio_galon.
        grupos_tmp: Dict[str, Dict] = {}
        for v in vehiculos:
            clave = v["tipo_combustible_clave"]
            if clave not in grupos_tmp:
                precio_key = _CLAVE_A_PRECIO.get(clave, "diesel")
                grupos_tmp[clave] = {
                    "label": _CLAVE_LABEL.get(clave, clave),
                    "precio_galon": self.precios.get(precio_key, PRECIOS_COMBUSTIBLE_DEFAULT["diesel"]),
                    "meses_usd": {},   # nombre_mes → total USD acumulado
                }
            for nombre_mes, val_mes in v["meses"].items():
                grupos_tmp[clave]["meses_usd"][nombre_mes] = round(
                    grupos_tmp[clave]["meses_usd"].get(nombre_mes, 0.0) + val_mes, 2
                )

        # Construir la estructura final por grupo con lista de meses ordenada
        grupos: Dict[str, Dict] = {}
        for clave, g in grupos_tmp.items():
            precio_galon = g["precio_galon"]
            meses_ordenados = sorted(
                g["meses_usd"].items(),
                key=lambda x: _mes_a_num.get(x[0], 99),
            )
            meses_list: List[Dict] = []
            for nombre_mes, total_mes_usd in meses_ordenados:
                galones_mes = round(total_mes_usd / precio_galon, 2) if precio_galon > 0 else 0.0
                meses_list.append({
                    "mes_numero": _mes_a_num.get(nombre_mes, 0),
                    "mes_nombre": nombre_mes,
                    "total_usd": total_mes_usd,
                    "galones": galones_mes,
                })
            total_anual_usd = round(sum(m["total_usd"] for m in meses_list), 2)
            total_anual_galones = round(sum(m["galones"] for m in meses_list), 2)
            grupos[clave] = {
                "label": g["label"],
                "precio_galon": precio_galon,
                "meses": meses_list,
                "total_anual_usd": total_anual_usd,
                "total_anual_galones": total_anual_galones,
            }

        print(
            f"✅ [CAT1] {len(vehiculos)} vehículos | "
            f"grupos={list(grupos.keys())}"
        )

        datos = self.initialize_data_dict()
        datos.update({
            "filename": filename,
            "anio": anio,
            "tipo_documento": "reporte_combustible",
            "grupos": grupos,
            "precios_galon": dict(self.precios),
            "extraction_success": True,
        })
        return datos

    # ── Extractor PDF (facturas individuales, otras zonas) ────────────────

    def _extract_pdf(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Extrae datos de una factura PDF individual de combustible."""
        try:
            texto = self.extract_pdf_text(file_content)
        except Exception as exc:
            return self.create_error_response(f"Error al leer el PDF: {exc}")

        if not texto or len(texto.strip()) < 20:
            return self.create_error_response("PDF sin texto legible.")

        datos = self.initialize_data_dict()
        datos["filename"] = filename
        datos["tipo_documento"] = "factura_combustible"

        # Número de factura  (ej. "No. 007-104-000153268")
        datos["nro_factura"] = (
            self.extract_string(
                r"(?:F\s*A\s*C\s*T\s*U\s*R\s*A|FACTURA)\s*\n?\s*No\.?\s*([\d\-]+)", texto
            )
            or self.extract_string(r"\bNo\.\s*([\d\-]+)", texto)
        )

        # Fecha de emisión  (ej. "Fecha Emisión: 20/01/2025")
        fecha_str = self.extract_string(
            r"Fecha\s+Emisi[oó]n\s*[:\s]+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})", texto
        )
        datos["fecha_emision"] = self.extract_date(fecha_str) if fecha_str else None

        # Detectar la línea de combustible en el detalle de la factura.
        # Para facturas con múltiples ítems (mantenimiento + combustible + transporte)
        # se extrae solo la fila que corresponde al combustible.
        cantidad, descripcion, precio_unit, total_linea = _detectar_linea_combustible(texto)
        datos["cantidad"] = cantidad
        datos["descripcion"] = descripcion
        datos["precio_unitario"] = precio_unit

        # total_usd: usar el total de la línea de combustible (no el VALOR A PAGAR total)
        if total_linea is not None and total_linea > 0:
            datos["total_usd"] = total_linea
        elif cantidad and precio_unit:
            datos["total_usd"] = round(cantidad * precio_unit, 2)
        else:
            datos["total_usd"] = self.extract_value(
                r"SUBTOTAL\s+SIN\s+IMPUESTOS\s+\$?([\d.,]+)", texto
            )

        # Subtotal sin impuestos de la factura completa (útil para facturas de un ítem)
        datos["subtotal"] = self.extract_value(
            r"SUBTOTAL\s+SIN\s+IMPUESTOS\s+\$?([\d.,]+)", texto
        )

        print(
            f"✅ [CAT1-PDF] nro={datos.get('nro_factura')} | "
            f"fecha={datos.get('fecha_emision')} | "
            f"desc={descripcion} | "
            f"cantidad={cantidad} | "
            f"total={datos.get('total_usd')}"
        )
        datos["extraction_success"] = True
        return datos
